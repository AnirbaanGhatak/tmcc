import pandas as pd
import openpyxl as opx

# # dataframe1 = pd.read_excel('hdfc.xlsx')
# # print(dataframe1)
# # dataframe1['Names'] = dataframe1['Names'].str.split('Credit Card').str[0] + "Credit Card"
# # print(dataframe1)

#     # df.to_excel("hdfcCCNames.xlsx",  f"{UrlName['Names'][index]}.xlsx", index=True)

# def get_maximum_rows(*, sheet_object):
#     rows = 0
#     for max_row, row in enumerate(sheet_object, 1):
#         if not all(col.value is None for col in row):
#             rows += 1
#     return rows

# workbook = opx.load_workbook('hdfcCCNames.xlsx')
# sheet_object = workbook.active
# max_rows = get_maximum_rows(sheet_object=sheet_object)

# print(max_rows)

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import openpyxl as opx
from openpyxl import *
# create a new Chrome browser instance

browser = webdriver.Chrome()

browser.get(f"https://cardinsider.com/hdfc-bank/")

cctrial = browser.find_elements(By.CLASS_NAME, "title_list_link")
ccls = [xs.text for xs in cctrial]
print(ccls) 

# df = pd.DataFrame({"Names": ccls})

# bankName = "hdfc-bank_CC"

# wb = Workbook()
# ws = wb.active
# wb.create_sheet("Sheet1")
# wb.save(filename = f"{bankName}_CC.xlsx")

# with pd.ExcelWriter(f"{bankName}_CC.xlsx", mode= 'a', engine= "openpyxl", if_sheet_exists='overlay') as writer1:
#     df.to_excel(writer1, sheet_name='Sheet1', index=False)

# close the browser
# browser.quit()

for x in ccls:
    


    # print(sfx)
    # print(x[sfx:])


